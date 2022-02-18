from os import path
from pprint import pprint
from traceback import print_exc

from openpyxl import load_workbook
from pandas import DataFrame, ExcelWriter

from indeed_crawler import IndeedCrawler

DEBUG = True


def append_df_to_excel(
        df, filename, sheet_name='Sheet1',
        **to_excel_kwargs
        ) -> None:
    if not path.exists(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            **to_excel_kwargs)
        return None
    writer = ExcelWriter(filename, engine='openpyxl', mode='a')
    writer.book = load_workbook(filename)
    start_row = writer.book[sheet_name].max_row
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    df.to_excel(writer, sheet_name, startrow=start_row,
                header=False, **to_excel_kwargs)
    writer.save()
    writer.close()
    return None


queries = [
]

places = [
    ('remote', 'united states')
]

negate_jobs = [
]

negate_comps = [
]

indeed_crawler = IndeedCrawler(
    number_of_jobs=0,
    debug=False,
    manually_fill_out_questions=False
    )
indeed_crawler.setup_browser()
indeed_crawler.login(
    email='',
    password=''
    )

i = 0
for region, country in places:
    for query in queries:
        try:
            indeed_crawler.search_jobs(
                query=query,
                job_title_negate_lst=negate_jobs,
                company_name_negate_lst=negate_comps,
                past_14_days=False,
                job_type='',  # fulltime
                min_salary='',
                enforce_salary=False,  # consider only jobs with salary listed
                exp_lvl='',  # entry_level, mid_level, #senior_level
                remote=remote,
                temp_remote=False,
                country=country,
                location=region,
                radius='',
                auto_answer_questions=True
                )

        except Exception as e:
            if DEBUG:
                print_exc()
            else:
                print(str(e))
        finally:
            results = indeed_crawler.results
            dataframe = DataFrame(data=results)
            print('\n\nApplied to {} jobs.'.format(len(dataframe)), end='\n\n')
            append_df_to_excel(dataframe, f'submissions{i}.xlsx',
                               sheet_name='job', index=False)
            i += 1
