from os import environ, path
from tkinter import Tk

from openpyxl import load_workbook
from pandas import DataFrame, ExcelWriter


def append_df_to_excel(
        df: DataFrame, filename: str, sheet_name='Sheet1',
        **to_excel_kwargs
        ) -> None:
    file_path = path.join(environ['USERPROFILE'], 'Desktop', filename)
    if not path.exists(file_path):
        df.to_excel(
            file_path,
            sheet_name=sheet_name,
            **to_excel_kwargs)
        return None
    writer = ExcelWriter(file_path, engine='openpyxl', mode='a')
    writer.book = load_workbook(file_path)
    start_row = writer.book[sheet_name].max_row
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    df.to_excel(writer, sheet_name, startrow=start_row,
                header=False, **to_excel_kwargs)
    writer.save()
    writer.close()
    return None


def float_convertible(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False


def int_convertible(s: str) -> bool:
    try:
        int(s)
        return True
    except ValueError:
        return False


def center(window: Tk) -> None:
    """
    centers a tkinter window
    :param window: the main window or Toplevel window to center
    """
    window.update_idletasks()
    width = window.winfo_width()
    frm_width = window.winfo_rootx() - window.winfo_x()
    window_width = width + 2 * frm_width
    height = window.winfo_height()
    titlebar_height = window.winfo_rooty() - window.winfo_y()
    window_height = height + titlebar_height + frm_width
    x = window.winfo_screenwidth() // 2 - window_width // 2
    y = window.winfo_screenheight() // 2 - window_height // 2
    window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    window.deiconify()
    return None
